# !/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Interactive multi dimensional molecular structure visualisation application
"""
import dash
import dash_bio as bio
import dash_core_components as dcc
import dash_html_components as html
import plotly.graph_objs as go
import pandas as pd
import random
from openpyxl import load_workbook
from utils import load_json

from plot import cluster_plot, val_plot, multi_plot

app = dash.Dash(__name__)
server = app.server

# Loading data table
df = pd.read_csv('./data/electronic_features.csv', index_col='ID')
df_q = load_workbook('./data/questionnaire.xlsx')
sheet = df_q.active

columns_dict = [{'label': i, 'value': i} for i in df.columns[3:15]]
size_dict = columns_dict + [{'label': 'Constant', 'value': 5}]

# Dash HTML application
app.layout = html.Div(
    id='main',
    className='app_main',
    children=[
        # Dash title and icon
        html.Div(
            id='mol3d-title',
            children=[
                html.Img(
                    id='dash-logo',
                    src="https://s3-us-west-1.amazonaws.com/plotly-tutorials/"
                        "logo/new-branding/dash-logo-by-plotly-stripe.png"
                ),
                html.H1(
                    id='chart-title',
                    children='Small-molecule photocatalysts'),
                html.Div(id='live-update-link')
            ]
        ),
        # Reload pages for selecting challenge randomly
        dcc.Interval(
                id='interval-component',
                interval=60*1000,  # in milliseconds
                n_intervals=0
        ),
        # Dash graph
        dcc.Graph(id='clickable_plot'),
        # Axes controller
        html.Div(
            id='plot-controller',
            children=[
                dcc.RadioItems(
                    id='chart_type',
                    options=[
                        {'label': '5D explorer', 'value': '5d'},
                        {'label': '2D chemical space', 'value': 'cluster'},
                        {'label': 'Blind tests', 'value': 'val'}
                    ],
                    value='cluster'
                ),
                html.Div(
                    id='color-bar-control',
                    className='dropdown-control',
                    children=[
                        html.H3('Symbol colour:'),
                        dcc.Dropdown(
                            id='colour_column',
                            options=columns_dict,
                            value='Hydrogen evolution rate / µmol/h'
                        )
                    ]
                ),
                html.Div(
                    id='size-control',
                    className='dropdown-control',
                    children=[
                        html.H3('Symbol size:'),
                        dcc.Dropdown(
                            id='size_column',
                            options=size_dict,
                            value=5
                        )
                    ]
                ),
                html.Div(
                    id='x-axis',
                    className='dropdown-control',
                    children=[
                        html.H3('X-axis:'),
                        dcc.Dropdown(
                            id='x_axis_column',
                            className='axis_controller',
                            options=columns_dict,
                            value='EA* / V'
                        )
                    ]
                ),
                html.Div(
                    id='y-axis',
                    className='dropdown-control',
                    children=[
                        html.H3('Y-axis:'),
                        dcc.Dropdown(
                            id='y_axis_column',
                            className='axis_controller',
                            options=columns_dict,
                            value='S1-T1 / eV'
                        )
                    ]
                ),
                html.Div(
                    id='z-axis',
                    className='dropdown-control',
                    children=[
                        html.H3('Z-axis:'),
                        dcc.Dropdown(
                            id='z_axis_column',
                            className='axis_controller',
                            options=columns_dict,
                            value='E(exciton binding) / eV'
                        )
                    ]
                ),
                html.P('Dropdown menus for symbol colour, symbol size, '
                       'X, Y, and Z axes are only active for 5D explorer.')
            ]
        ),
        # Molecule 3D viewer
        html.Div(
            id='selected_structure',
            children=[
                html.H3(className='viewer-title', children='Selected structure:'),
                dcc.Loading(id='loading_selected', className='loading')
            ]
        ),
    ]
)


def structure_viewer(interactive_data, chart_name):

    def single_3d_viewer(json_file, structure_index):
        mol_data, style_data = load_json(json_file)
        mol_3d = html.Div(
            id='viewer',
            children=[
                html.P('Structure ID: {}'.format(structure_index)),
                bio.Molecule3dViewer(
                    id='mol-3d-viewer',
                    selectionType='atom',
                    styles=style_data,
                    modelData=mol_data
                )
            ]
        )
        return mol_3d
    mol_div = []
    # Extracting molecular index from Dash interactive data
    try:
        for i in range(len(interactive_data['points'])):
            # 5D figure
            if chart_name == '5d':
                index = int(interactive_data['points'][0]['pointNumber'])
                structure_name = df.iloc[index].name
                id_in_paper = df.iloc[index, 0]
            # Validation figure
            elif chart_name == 'val':
                index = int(interactive_data['points'][i]['pointIndex'])
                cluster_idx = interactive_data['points'][i]['curveNumber']
                if cluster_idx == 0:
                    structure_name = df[df.index <= 816].iloc[index].name
                    id_in_paper = df[df.index <= 816].iloc[index][0]
                else:
                    structure_name = df[df.index > 816].iloc[index].name - 1
                    id_in_paper = df[df.index > 816].iloc[index][0] - 1
            # Cluster figure
            else:
                index = int(interactive_data['points'][i]['pointIndex'])
                cluster_idx = interactive_data['points'][i]['curveNumber'] + 1
                structure_name = df[df.group == cluster_idx].iloc[index].name
                id_in_paper = df[df.group == cluster_idx].iloc[index][0]
            json_path = './data/json_data/{}.json'.format(structure_name)
            # Add molecular viewer
            viewer = single_3d_viewer(json_path, int(id_in_paper))
            mol_div.append(viewer)
    # Default structure
    except TypeError:
        json_path = './data/json_data/340.json'
        mol_div.append(single_3d_viewer(json_path, 'default 153'))
    return mol_div


@app.callback(
    dash.dependencies.Output('clickable_plot', 'figure'),
    [dash.dependencies.Input('chart_type', 'value'),
     dash.dependencies.Input('x_axis_column', 'value'),
     dash.dependencies.Input('y_axis_column', 'value'),
     dash.dependencies.Input('z_axis_column', 'value'),
     dash.dependencies.Input('colour_column', 'value'),
     dash.dependencies.Input('size_column', 'value')]
)
def update_graph(chart_type_value, x_axis_column_name, y_axis_column_name,
                 z_axis_column_name, colour_column_value, size_column_value):
    fig = go.Figure()
    # Update figures by different type name
    if chart_type_value == 'cluster':
        fig = cluster_plot(fig, df)
    elif chart_type_value == '5d':
        fig = multi_plot(
            fig, df, x_axis_column_name, y_axis_column_name,
            z_axis_column_name, size_column_value, colour_column_value
        )
    elif chart_type_value == 'val':
        fig = val_plot(fig, df)
    return fig


@app.callback(dash.dependencies.Output('loading_selected', 'children'),
              [dash.dependencies.Input('clickable_plot', 'clickData'),
               dash.dependencies.Input('clickable_plot', 'selectedData'),
               dash.dependencies.Input('chart_type', 'value')])
def display_selected_data(clickData, selectedData, chart_type_value):
    if chart_type_value == '5d':
        data = clickData
    else:
        data = selectedData
    return structure_viewer(interactive_data=data, chart_name=chart_type_value)


@app.callback(dash.dependencies.Output('live-update-link', 'children'),
              [dash.dependencies.Input('interval-component', 'n_intervals'),
               dash.dependencies.Input('chart_type', 'value')])
def update_link(n, chart_type_value):
    return html.A(
        "Take the HER prediction challenges",
        href=sheet.cell(row=random.randint(1, 20), column=2).value
    )


if __name__ == '__main__':
    app.run_server(debug=True, use_reloader=True)
